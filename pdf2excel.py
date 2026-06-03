# -*- coding: utf-8 -*-
"""
PDF 批量转 Excel（一个 PDF 一个工作表）。

重点：尽量还原 PDF 里的表格显示格式（行列、合并单元格、边框），
同时保证不丢内容。

识别思路：
1. PyMuPDF 把每页渲染成图。
2. RapidOCR 识别文字（得到带坐标的文本框）。
3. 【主路】reconstruct_from_ocr：直接用 OCR 文本框坐标重建表格——
   用“框最多”的一行定表头与列边界，用最左列（序号）定逻辑行边界，
   每个文本框按中心点归到对应单元格。多行折行的账号/名称不会错位，
   且不依赖表格模型那套不稳定的内部文本归位。
4. 回退：表格结构模型（wired/lineless）→ 坐标启发式拼表。
5. table_to_excel 把单元格写成 Excel，还原合并单元格和边框。
6. 把表格没接住的 OCR 文字补在表下面（保证不丢）；低置信度文字（水印）
   不进表格，只进兜底区。

环境：Windows + VSCode，只能 pip 装包，首次联网下模型，之后可离线。
"""
from __future__ import annotations

import argparse
import importlib
import os
import re
from dataclasses import dataclass
from typing import List, Optional, Tuple

import numpy as np


# ----------------------------------------------------------------------------
# 渲染 PDF
# ----------------------------------------------------------------------------
def render_pdf_pages(path: str, dpi: int = 300):
    import fitz  # PyMuPDF

    doc = fitz.open(path)
    total = doc.page_count
    scale = dpi / 72.0
    for i in range(total):
        page = doc[i]
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
        img = np.frombuffer(pix.samples, np.uint8).reshape(pix.height, pix.width, pix.n)
        if pix.n == 4:
            img = img[:, :, :3]
        elif pix.n == 1:
            img = np.repeat(img, 3, axis=2)
        yield i, total, img[:, :, ::-1].copy()  # BGR
    doc.close()


# ----------------------------------------------------------------------------
# OCR（RapidOCR）
# ----------------------------------------------------------------------------
_OCR_ENGINE = None


def get_ocr_engine():
    global _OCR_ENGINE
    if _OCR_ENGINE is None:
        from rapidocr_onnxruntime import RapidOCR
        _OCR_ENGINE = RapidOCR()
    return _OCR_ENGINE


def ocr_image(img_bgr) -> list:
    """返回 RapidOCR 格式：[[box(4点), text, score], ...]，没识别到返回 []。"""
    engine = get_ocr_engine()
    result, _ = engine(img_bgr)
    return result or []


# ----------------------------------------------------------------------------
# 文本框
# ----------------------------------------------------------------------------
@dataclass
class Box:
    text: str
    x0: float
    y0: float
    x1: float
    y1: float

    @property
    def cx(self):
        return (self.x0 + self.x1) / 2

    @property
    def cy(self):
        return (self.y0 + self.y1) / 2

    @property
    def h(self):
        return self.y1 - self.y0


def _to_boxes(ocr_result: list) -> List[Box]:
    boxes = []
    for item in ocr_result:
        try:
            quad, text = item[0], item[1]
            xs = [p[0] for p in quad]
            ys = [p[1] for p in quad]
            boxes.append(Box(str(text), min(xs), min(ys), max(xs), max(ys)))
        except Exception:
            continue
    return boxes


def _norm(s) -> str:
    return re.sub(r"\s+", "", str(s))


# ----------------------------------------------------------------------------
# 表格结构识别引擎
# 重要：WiredTableRecognition / LinelessTableRecognition 不是无参构造，
# 需要传 WiredTableInput / LinelessTableInput（配置类常在 .main 里，不一定从包根导出）。
# ----------------------------------------------------------------------------
_WIRED = None
_LINELESS = None


def _build_engine(kind: str):
    if kind == "wired":
        pkg_name, rec_name, input_name = "wired_table_rec", "WiredTableRecognition", "WiredTableInput"
    else:
        pkg_name, rec_name, input_name = "lineless_table_rec", "LinelessTableRecognition", "LinelessTableInput"

    pkg = importlib.import_module(pkg_name)
    rec_cls = getattr(pkg, rec_name, None)
    input_cls = getattr(pkg, input_name, None)

    if rec_cls is None or input_cls is None:
        try:
            main_mod = importlib.import_module(pkg_name + ".main")
            rec_cls = rec_cls or getattr(main_mod, rec_name, None)
            input_cls = input_cls or getattr(main_mod, input_name, None)
        except Exception:
            pass

    if rec_cls is None:
        raise ImportError(f"找不到 {rec_name}，请确认已 pip install {pkg_name}")

    last_err = None
    if input_cls is not None:
        try:
            return rec_cls(input_cls())
        except Exception as e:
            last_err = e
    try:
        return rec_cls()
    except Exception as e:
        last_err = e
    raise last_err


def _load_wired():
    global _WIRED
    if _WIRED is None:
        _WIRED = _build_engine("wired")
    return _WIRED


def _load_lineless():
    global _LINELESS
    if _LINELESS is None:
        _LINELESS = _build_engine("lineless")
    return _LINELESS


def _to_xyxy(bbox) -> Optional[Tuple[float, float, float, float]]:
    """把 cell bbox 统一成 (x0,y0,x1,y1)。支持 [x0,y0,x1,y1]、
    8 个数的多边形、或 4 个点的多边形。"""
    if bbox is None:
        return None
    arr = np.asarray(bbox, dtype=float).flatten()
    if arr.size == 4:
        x0, y0, x1, y1 = arr
        return (min(x0, x1), min(y0, y1), max(x0, x1), max(y0, y1))
    if arr.size >= 8 and arr.size % 2 == 0:
        xs = arr[0::2]; ys = arr[1::2]
        return (float(xs.min()), float(ys.min()), float(xs.max()), float(ys.max()))
    return None


def _extract_struct(ret):
    """从不同版本的返回值里拿出 (html, cell_bboxes, logic_points)。
    任何一项拿不到就是 None。"""
    html = getattr(ret, "pred_html", None)
    cell_bboxes = getattr(ret, "cell_bboxes", None)
    logic_points = getattr(ret, "logic_points", None)

    if isinstance(ret, (tuple, list)):
        if html is None and ret and isinstance(ret[0], str) and "<" in ret[0]:
            html = ret[0]
        if cell_bboxes is None and len(ret) >= 3:
            cell_bboxes = ret[2]
        if logic_points is None and len(ret) >= 4:
            logic_points = ret[3]
    elif isinstance(ret, str) and "<" in ret:
        html = html or ret

    return html, cell_bboxes, logic_points


def _group_lines(items: List[Box]) -> str:
    """把一个单元格里的多个文本框按阅读顺序（从上到下、从左到右）拼接。
    中文/数字场景直接拼接，不加空格（账号换行合并不会丢位）。"""
    if not items:
        return ""
    med_h = float(np.median([b.h for b in items if b.h > 0]) or 12)
    items = sorted(items, key=lambda b: b.cy)
    lines: List[List[Box]] = [[items[0]]]
    for b in items[1:]:
        if abs(b.cy - lines[-1][-1].cy) <= med_h * 0.6:
            lines[-1].append(b)
        else:
            lines.append([b])
    parts = []
    for line in lines:
        line.sort(key=lambda b: b.x0)
        parts.append("".join(b.text for b in line))
    return "".join(parts)


def build_cells_by_geometry(cell_bboxes, logic_points, ocr_result: list):
    """用单元格几何 + OCR 文本框中心包含关系，自己做文本归位。
    返回 (cells, n_rows, n_cols, covered_keys) 或 None（几何信息不可用）。
    cells: [{r0,r1,c0,c1,text}]
    """
    if cell_bboxes is None or logic_points is None:
        return None
    try:
        bboxes = [_to_xyxy(b) for b in cell_bboxes]
        lps = [list(np.asarray(p, dtype=float).flatten()) for p in logic_points]
    except Exception:
        return None
    if not bboxes or len(bboxes) != len(lps):
        return None

    boxes = _to_boxes(ocr_result)
    cell_items: List[List[Box]] = [[] for _ in bboxes]
    covered_keys = set()

    def center_in(bb, b):
        x0, y0, x1, y1 = bb
        pad = max(2.0, (y1 - y0) * 0.15)  # 略微外扩，容忍单元格边界偏差
        return (x0 - pad) <= b.cx <= (x1 + pad) and (y0 - pad) <= b.cy <= (y1 + pad)

    for b in boxes:
        best, best_area = -1, -1.0
        for i, bb in enumerate(bboxes):
            if bb is None:
                continue
            if center_in(bb, b):
                area = (bb[2] - bb[0]) * (bb[3] - bb[1])
                if best == -1 or area < best_area:  # 嵌套时选最小单元格
                    best, best_area = i, area
        if best >= 0:
            cell_items[best].append(b)
            covered_keys.add(_norm(b.text))

    cells = []
    max_r = max_c = 0
    for i, lp in enumerate(lps):
        if len(lp) < 4:
            continue
        r0, r1, c0, c1 = int(lp[0]), int(lp[1]), int(lp[2]), int(lp[3])
        cells.append({"r0": r0, "r1": r1, "c0": c0, "c1": c1, "text": _group_lines(cell_items[i])})
        max_r = max(max_r, r1); max_c = max(max_c, c1)
    if not cells:
        return None
    return cells, max_r + 1, max_c + 1, covered_keys


def recognize_table(img_bgr, ocr_result: list, prefer: str = "wired"):
    """返回一个 dict：
      {"type":"struct", "cells":..., "nrows":..., "ncols":..., "covered":set}  最佳
      {"type":"html",   "html":...}                                          次优
    都失败返回 None。"""
    order = ["wired", "lineless"] if prefer == "wired" else ["lineless", "wired"]
    fallback_html = None
    for kind in order:
        try:
            engine = _load_wired() if kind == "wired" else _load_lineless()
        except Exception as e:
            print(f"    [表格识别-{kind}] 加载失败，跳过：{e}")
            continue
        try:
            try:
                ret = engine(img_bgr, ocr_result=ocr_result)
            except TypeError:
                ret = engine(img_bgr)
            html, cell_bboxes, logic_points = _extract_struct(ret)
            built = build_cells_by_geometry(cell_bboxes, logic_points, ocr_result)
            if built:
                cells, nrows, ncols, covered = built
                return {"type": "struct", "cells": cells, "nrows": nrows,
                        "ncols": ncols, "covered": covered}
            if html and "<td" in html.lower() and fallback_html is None:
                fallback_html = html
        except Exception as e:
            print(f"    [表格识别-{kind}] 运行出错，跳过：{e}")
    if fallback_html:
        return {"type": "html", "html": fallback_html}
    return None


# ----------------------------------------------------------------------------
# 不丢内容：找出表格没接住的 OCR 文字
# ----------------------------------------------------------------------------
def find_missing_texts(ocr_result: list, covered_text: str = "", covered_keys=None) -> List[str]:
    """返回 OCR 识别到、但没被表格接住的文字（去重后）。
    - struct 模式：传 covered_keys（已归位的 _norm 文本集合）
    - html 模式：传 covered_text（表格纯文本）
    """
    covered_keys = covered_keys or set()
    table_text = _norm(covered_text)
    missing, seen = [], set()
    for item in ocr_result:
        if len(item) < 2:
            continue
        raw = str(item[1]).strip()
        key = _norm(raw)
        if not key or key in seen:
            continue
        if key in covered_keys:
            continue
        if table_text and key in table_text:
            continue
        missing.append(raw)
        seen.add(key)
    return missing


# ----------------------------------------------------------------------------
# 回退方案：按坐标启发式拼表（无线 / 模型不可用时）
# ----------------------------------------------------------------------------
def reconstruct_grid(ocr_result: list) -> List[List[str]]:
    """按行聚类 + 列聚类的启发式表格（不丢文字）。"""
    boxes = _to_boxes(ocr_result)
    if not boxes:
        return []
    boxes.sort(key=lambda b: b.cy)
    med_h = float(np.median([b.h for b in boxes if b.h > 0]) or 12)

    rows: List[List[Box]] = []
    cur: List[Box] = [boxes[0]]
    for b in boxes[1:]:
        if abs(b.cy - cur[-1].cy) <= med_h * 0.7:
            cur.append(b)
        else:
            rows.append(cur); cur = [b]
    rows.append(cur)

    lefts = sorted(b.x0 for b in boxes)
    cols: List[float] = [lefts[0]]
    for x in lefts[1:]:
        if x - cols[-1] > med_h * 1.5:
            cols.append(x)
    ncols = len(cols)

    def col_of(b: Box) -> int:
        return min(range(ncols), key=lambda i: abs(b.x0 - cols[i]))

    grid: List[List[str]] = []
    for row in rows:
        line = [""] * ncols
        for b in sorted(row, key=lambda b: b.x0):
            ci = col_of(b)
            line[ci] = (line[ci] + " " + b.text).strip() if line[ci] else b.text
        grid.append(line)
    return grid


# ----------------------------------------------------------------------------
# 写 Excel
# ----------------------------------------------------------------------------
def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/?*\[\]:]", "_", name)
    name = name.strip() or "Sheet"
    return name[:31]


def _thin_border():
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_grid_with_borders(ws, grid: List[List[str]], start_row: int = 1) -> int:
    from openpyxl.styles import Alignment
    border = _thin_border()
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ncols = max((len(r) for r in grid), default=0)
    for ri, row in enumerate(grid):
        for ci in range(ncols):
            val = row[ci] if ci < len(row) else ""
            cell = ws.cell(start_row + ri, ci + 1, val if val else None)
            cell.border = border
            cell.alignment = align
    return start_row + len(grid)


def _write_missing_block(ws, missing: List[str], start_row: int) -> int:
    if not missing:
        return start_row
    from openpyxl.styles import Font
    r = start_row + 1
    c = ws.cell(r, 1, "↓ 以下为未归入表格的识别文字（请人工核对）")
    c.font = Font(bold=True, color="C00000")
    r += 1
    for t in missing:
        ws.cell(r, 1, t)
        r += 1
    return r


def find_pdfs(input_path: str) -> List[str]:
    if os.path.isfile(input_path) and input_path.lower().endswith(".pdf"):
        return [input_path]
    out = []
    for root, _, files in os.walk(input_path):
        for f in files:
            if f.lower().endswith(".pdf"):
                out.append(os.path.join(root, f))
    return sorted(out)


def _ocr_to_debug(ocr_result: list):
    out = []
    for item in ocr_result:
        try:
            quad, text = item[0], item[1]
            score = float(item[2]) if len(item) > 2 else 1.0
            xs = [p[0] for p in quad]
            ys = [p[1] for p in quad]
            out.append({"text": str(text), "score": round(score, 3),
                        "x0": round(min(xs), 1), "y0": round(min(ys), 1),
                        "x1": round(max(xs), 1), "y1": round(max(ys), 1)})
        except Exception:
            continue
    return out


def convert(input_path: str, output_path: str, dpi: int = 300,
            min_score: float = 0.5, debug: bool = False):
    import json
    from openpyxl import Workbook
    from table_to_excel import write_table_to_sheet, write_cells_to_sheet
    from bs4 import BeautifulSoup
    from reconstruct import reconstruct_from_ocr

    debug_pages = []

    pdfs = find_pdfs(input_path)
    if not pdfs:
        print(f"没找到 PDF：{input_path}")
        return

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    for pdf in pdfs:
        base = os.path.splitext(os.path.basename(pdf))[0]
        name = sanitize_sheet_name(base)
        n, k = name, 1
        while n in used_names:
            k += 1
            n = sanitize_sheet_name(f"{base}_{k}")
        used_names.add(n)
        ws = wb.create_sheet(n)
        print(f"\n处理：{pdf} -> 工作表 [{n}]")

        row_cursor = 1
        for pno, total, img in render_pdf_pages(pdf, dpi=dpi):
            print(f"  第 {pno + 1}/{total} 页 OCR...")
            ocr_result = ocr_image(img)
            if not ocr_result:
                print("    （未识别到文字）")
                continue

            dbg = {} if debug else None
            recon = reconstruct_from_ocr(ocr_result, min_score=min_score, debug=dbg)
            if debug:
                debug_pages.append({
                    "pdf": os.path.basename(pdf), "page": pno + 1,
                    "ocr": _ocr_to_debug(ocr_result), "recon": dbg,
                })

            if recon is not None:
                cells, nrows, ncols, covered = recon
                missing = find_missing_texts(ocr_result, covered_keys=covered)
                print(f"    ✓ 按坐标重建表格（{nrows}行×{ncols}列）；未归入文字 {len(missing)} 条")
                row_cursor = write_cells_to_sheet(ws, cells, nrows, ncols, start_row=row_cursor)
                row_cursor = _write_missing_block(ws, missing, row_cursor)
                row_cursor += 1
                continue

            # 回退：表格结构模型
            res = recognize_table(img, ocr_result, prefer="wired")
            if res and res["type"] == "struct":
                missing = find_missing_texts(ocr_result, covered_keys=res["covered"])
                print(f"    ✓ 表格识别成功（模型几何归位）；未归入文字 {len(missing)} 条")
                row_cursor = write_cells_to_sheet(ws, res["cells"], res["nrows"], res["ncols"], start_row=row_cursor)
                row_cursor = _write_missing_block(ws, missing, row_cursor)
                row_cursor += 1
            elif res and res["type"] == "html":
                table_text = BeautifulSoup(res["html"], "lxml").get_text()
                missing = find_missing_texts(ocr_result, covered_text=table_text)
                print(f"    ✓ 表格识别成功（HTML 回退）；未归入文字 {len(missing)} 条")
                row_cursor = write_table_to_sheet(ws, res["html"], start_row=row_cursor)
                row_cursor = _write_missing_block(ws, missing, row_cursor)
                row_cursor += 1
            else:
                print("    · 未检测到表格，回退到坐标拼表")
                grid = reconstruct_grid(ocr_result)
                if grid:
                    row_cursor = _write_grid_with_borders(ws, grid, start_row=row_cursor) + 1

        for col in ws.columns:
            try:
                letter = col[0].column_letter
            except Exception:
                continue
            maxlen = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[letter].width = min(40, max(8, maxlen + 2))

    wb.save(output_path)
    print(f"\n完成：{output_path}")

    if debug:
        dbg_path = output_path + ".debug.json"
        with open(dbg_path, "w", encoding="utf-8") as f:
            json.dump(debug_pages, f, ensure_ascii=False, indent=2)
        print(f"调试信息已写入：{dbg_path}（若仍有错位，请把这个文件发我）")


def main():
    ap = argparse.ArgumentParser(description="PDF 批量转 Excel（还原表格格式，不丢内容）")
    ap.add_argument("-i", "--input", required=True, help="PDF 文件或目录")
    ap.add_argument("-o", "--output", default="out.xlsx", help="输出 xlsx")
    ap.add_argument("--dpi", type=int, default=300, help="渲染 DPI（默认 300，扫描件不清晰可调 400）")
    ap.add_argument("--min-score", type=float, default=0.5,
                    help="OCR 置信度阈值（默认 0.5），低于此值的文字不进表格、只放到表下兜底区，可过滤水印")
    ap.add_argument("--debug", action="store_true",
                    help="导出每页 OCR 坐标与重建结果到 <输出>.debug.json，便于排查错位")
    args = ap.parse_args()
    convert(args.input, args.output, dpi=args.dpi, min_score=args.min_score, debug=args.debug)


if __name__ == "__main__":
    main()
